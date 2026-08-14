import copy
import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "fisa_accesorii_demo.xlsx"

FIRST_DATA_ROW = 6
HEADER_ROW = 5


def _copy_row_style(ws, src_row, dst_row, ncols=5):
    """Copiază stilul unui rând de articole pe un rând nou."""
    for col in range(1, ncols + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)

        if src.has_style:
            dst._style = copy.copy(src._style)

        if src.number_format:
            dst.number_format = src.number_format

        if src.alignment:
            dst.alignment = copy.copy(src.alignment)

        if src.font:
            dst.font = copy.copy(src.font)

        if src.border:
            dst.border = copy.copy(src.border)

        if src.fill:
            dst.fill = copy.copy(src.fill)

        if src.protection:
            dst.protection = copy.copy(src.protection)

    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def build_fisa(
    items,
    titlu,
    proiect,
    data,
    nr_document,
    fisa_de=None,
    depozit=None,
    output_path="fisa_accesorii.xlsx",
    template_path=None,
):
    """
    Generează fișa Excel pornind de la template.

    Coloanele D (Conf) și E (Dif) rămân libere pentru completare manuală.
    """

    template = Path(template_path) if template_path else TEMPLATE_PATH

    if not template.exists():
        raise FileNotFoundError(
            f"Template-ul Excel nu a fost găsit: {template}"
        )

    wb = load_workbook(template)
    ws = wb["Sheet1"]

    # Găsește ultimul rând de articole din template.
    template_last_data_row = None

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        value = ws.cell(row=r, column=1).value

        if value not in (None, "") and isinstance(value, (int, float)):
            template_last_data_row = r
        else:
            break

    if template_last_data_row is None:
        raise ValueError(
            "Nu s-a putut identifica zona de articole din template."
        )

    footer_row_1 = template_last_data_row + 1
    footer_row_2 = footer_row_1 + 1

    # Eliminăm merge-urile din footer înainte de modificarea numărului de rânduri.
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= footer_row_1:
            ws.unmerge_cells(str(rng))

    n_items = len(items)
    n_template_rows = template_last_data_row - FIRST_DATA_ROW + 1

    if n_items == 0:
        raise ValueError("Nu există articole de introdus în fișă.")

    # Ajustează numărul de rânduri din tabel.
    if n_items > n_template_rows:
        rows_to_add = n_items - n_template_rows
        ws.insert_rows(template_last_data_row + 1, rows_to_add)

        # Copiem stilul ultimului rând existent de articole.
        style_source_row = FIRST_DATA_ROW

        for i in range(n_template_rows, n_items):
            _copy_row_style(
                ws,
                style_source_row,
                FIRST_DATA_ROW + i,
            )

    elif n_items < n_template_rows:
        rows_to_delete = n_template_rows - n_items
        ws.delete_rows(
            FIRST_DATA_ROW + n_items,
            rows_to_delete,
        )

    footer_row_1 = FIRST_DATA_ROW + n_items
    footer_row_2 = footer_row_1 + 1

    # Footer-ul original folosește C:E unite pe două rânduri.
    ws.merge_cells(f"C{footer_row_1}:E{footer_row_2}")

    # --- Header ---
    ws["A3"] = titlu

    if isinstance(data, datetime.datetime):
        ws["C3"] = data
    elif isinstance(data, datetime.date):
        ws["C3"] = data
    elif data:
        try:
            ws["C3"] = datetime.datetime.strptime(
                str(data),
                "%d/%m/%Y",
            )
        except ValueError:
            ws["C3"] = data
    else:
        ws["C3"] = None

    ws["A4"] = proiect
    ws["C4"] = nr_document

    # --- Rânduri articole ---
    for i, item in enumerate(items):
        r = FIRST_DATA_ROW + i

        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=item["denumire"])
        ws.cell(row=r, column=3, value=item["cantitate"])

        # D = Conf și E = Dif rămân libere.

    # --- Footer ---
    if fisa_de is None:
        footer_1 = "Fisa Accesorii :   "
    else:
        footer_1 = f"Fisa Accesorii :   {fisa_de}"

    if depozit is None:
        footer_2 = "Depozit       :        "
    else:
        footer_2 = f"Depozit       :        {depozit}"

    ws.cell(row=footer_row_1, column=1, value=footer_1)
    ws.cell(row=footer_row_1, column=3, value=datetime.date.today())
    ws.cell(row=footer_row_2, column=1, value=footer_2)

    # Elimină eventualele rânduri rămase după footer.
    if ws.max_row > footer_row_2:
        ws.delete_rows(
            footer_row_2 + 1,
            ws.max_row - footer_row_2,
        )

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output)

    return str(output)


if __name__ == "__main__":
    from extract_pdf import extract_proforma

    pdf_path = "FP_260800280.pdf"

    data = extract_proforma(pdf_path)

    linia = data["proiect"] or ""

    if " - " in linia:
        titlu, proiect = [
            p.strip()
            for p in linia.split(" - ", 1)
        ]
    else:
        titlu = linia
        proiect = ""

    build_fisa(
        items=data["items"],
        titlu=titlu,
        proiect=proiect,
        data=data["data_livrare"],
        nr_document=data["nr_intern"],
        output_path="fisa_accesorii_demo_output.xlsx",
    )

    print("done")
